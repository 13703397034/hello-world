#记录excle openpyxl包的操作 https://www.jianshu.com/p/3f348b7552a7
from openpyxl import Workbook
#导入包注意大小写，定义一个workbook类
wb = Workbook()
#定位到工作表sheet(初始化一个sheet实例)
ws = wb.active
# idex 等同于索引标记
ws1 = wb.create_sheet("mysheet1")# 插入到最后
ws2 = wb.create_sheet("mysheet2",0)# 插入到最前面
ws3 = wb.create_sheet("mysheet3",-1)# 插入到倒数第二个
ws1.title="firstsheet"#修改sheet名称
ws.sheet_properties.tabColor="1072BA"#可以修改标签背景色为RRGGBB格式的颜色
#ws3 = wb["newtitle"]#可以将名称作为工作表的索引
print(wb.sheetnames)# 打印所有sheet名
for sheet in wb:      #遍历整个工作簿
    print(sheet.title)
source = wb.active# 重新定义一个sheet实例化
source1 = wb.copy_worksheet(ws2)#复制ws1到source1中
print(source1.title)
'''
以上内容为操作工作簿(sheet)的操作
'''
c = ws['A5']
ws['A5'] = 10000
d = ws.cell(6,6,20000)
print(d.coordinate)

